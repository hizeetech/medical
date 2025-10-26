from functools import lru_cache
from typing import Dict, List, Tuple, Optional, Any

from .models import FacilityExcelUpload

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None

STATE_CHOICES = [('OGS', 'Ogun State')]
LGA_NUMBER_CHOICES = [(f"{i:02d}", f"{i:02d}") for i in range(1, 41)]
FACILITY_TYPE_CHOICES = [(str(i), str(i)) for i in (1, 2, 3)]
FACILITY_NUMBER_CHOICES = [(f"{i:04d}", f"{i:04d}") for i in range(1, 601)]


def _normalize_num_str(val: Any, width: Optional[int] = None) -> str:
    if val is None:
        return ''
    try:
        if isinstance(val, (int,)):
            n = int(val)
        elif isinstance(val, (float,)):
            n = int(val)
        else:
            s = str(val).strip()
            if s.isdigit():
                n = int(s)
            else:
                return s.strip()
        return f"{n:0{width}d}" if width else str(n)
    except Exception:
        return str(val).strip()


@lru_cache(maxsize=1)
def get_facility_data() -> Dict[str, object]:
    upload = FacilityExcelUpload.objects.order_by('-uploaded_at').first()
    lga_set = []
    abbr_map: Dict[str, str] = {}
    if not upload or not upload.file or openpyxl is None:
        return {
            'lga_choices': [],
            'abbr_map': {},
        }

    wb = openpyxl.load_workbook(upload.file.path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=4):
        name_cell = row[0].value  # Column A: LGA name (typical layout)
        abbr_cell = row[1].value  # Column B: LGA abbreviation
        name = str(name_cell).strip() if name_cell else ''
        abbr = str(abbr_cell).strip() if abbr_cell else ''
        if not name or not abbr:
            continue
        if name not in abbr_map:
            lga_set.append(name)
        abbr_map[name] = abbr
    lga_choices = [(name, name) for name in sorted(lga_set)]
    return {
        'lga_choices': lga_choices,
        'abbr_map': abbr_map,
    }


def make_hospital_clinic_id(state_code: str, lga_name: str, lga_number: str, facility_type: str, facility_number: str) -> str:
    data = get_facility_data()
    abbr_map: Dict[str, str] = data.get('abbr_map', {})  # type: ignore
    lga_abbr = abbr_map.get(lga_name, '').upper()
    return f"{state_code}/PHCDB/{lga_abbr}/{lga_number}/{facility_type}/{facility_number}"


def parse_hospital_clinic_id(hospital_id: str) -> Optional[Tuple[str, str, str, str, str]]:
    if not hospital_id:
        return None
    parts = hospital_id.split('/')
    if len(parts) < 6:
        return None
    state, phcdb, lga_abbr, lga_number, facility_type, facility_number = parts[:6]
    return state, lga_abbr, lga_number, facility_type, facility_number


@lru_cache(maxsize=1024)
def find_facility_name(lga_name: Optional[str] = None, lga_number: Optional[str] = None, facility_type: Optional[str] = None, facility_number: Optional[str] = None, lga_abbr: Optional[str] = None) -> Optional[str]:
    upload = FacilityExcelUpload.objects.order_by('-uploaded_at').first()
    if not upload or not upload.file or openpyxl is None:
        return None

    wb = openpyxl.load_workbook(upload.file.path, read_only=True, data_only=True)
    ws = wb.active

    # Discover header indices dynamically (works with both older simple layout and the CODE block layout)
    header_row_idx = None
    header_indices: Dict[str, List[int]] = {
        'LGAABBREVIATION': [],
        'NAMEOFHEALTHFACILITY': [],
        'LGANUMBER': [],  # textual header 'LGA NUMBER'
        'LGA': [],        # numeric code column under CODE
        'FACILITYTYPE': [],  # both textual and numeric; we'll pick numeric
        'FACILITYNO': [],
    }

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
        header_hit = False
        for j, cell in enumerate(row):
            raw = str(cell.value).strip().upper() if cell.value else ''
            norm = raw.replace(' ', '')
            if not raw:
                continue
            if 'NAME' in raw and 'FACILITY' in raw and 'HEALTH' in raw:
                header_indices['NAMEOFHEALTHFACILITY'].append(j)
                header_row_idx = header_row_idx or i
                header_hit = True
            elif 'LGA' in raw and 'ABBREVIATION' in raw:
                header_indices['LGAABBREVIATION'].append(j)
                header_row_idx = header_row_idx or i
                header_hit = True
            elif raw == 'LGA NUMBER' or norm == 'LGANUMBER':
                header_indices['LGANUMBER'].append(j)
                header_row_idx = header_row_idx or i
                header_hit = True
            elif raw == 'LGA':
                header_indices['LGA'].append(j)
                header_row_idx = header_row_idx or i
                header_hit = True
            elif raw == 'FACILITY TYPE' or norm == 'FACILITYTYPE':
                header_indices['FACILITYTYPE'].append(j)
                header_row_idx = header_row_idx or i
                header_hit = True
            elif raw == 'FACILITY NO' or norm in ('FACILITYNO', 'FACILITYNUMBER'):
                header_indices['FACILITYNO'].append(j)
                header_row_idx = header_row_idx or i
                header_hit = True
        # use the first row that looks like a header containing name/abbr as the start
        if header_hit and header_row_idx:
            # don't break immediately; we want to collect indices across up to 20 rows to capture multi-row header blocks
            pass

    start_row = (header_row_idx or 2) + 1

    # Targets from input
    tgt_name = (lga_name or '').strip().upper()
    tgt_abbr = (lga_abbr or '').strip().upper()
    tgt_num = _normalize_num_str(lga_number, 2)
    tgt_type = _normalize_num_str(facility_type, None)
    tgt_fac = _normalize_num_str(facility_number, 4)

    fallback_match_name: Optional[str] = None

    for row in ws.iter_rows(min_row=start_row):
        # Pull columns by discovered indices
        def cell(idx_list: List[int]) -> Optional[Any]:
            for idx in idx_list:
                if idx < len(row):
                    v = row[idx].value
                    if v is not None and str(v).strip() != '':
                        return v
            return None

        abbr_val = cell(header_indices['LGAABBREVIATION'])
        name_val = cell(header_indices['NAMEOFHEALTHFACILITY'])
        lga_number_textual = cell(header_indices['LGANUMBER'])

        # choose numeric LGA code column: prefer an index whose cell is numeric
        lga_number_numeric = None
        for idx in header_indices['LGA']:
            if idx < len(row):
                v = row[idx].value
                s = str(v).strip() if v is not None else ''
                if s and s.replace('.', '', 1).isdigit():
                    lga_number_numeric = v
                    break

        # choose facility type numeric code when available
        facility_type_numeric = None
        facility_type_text = None
        for idx in header_indices['FACILITYTYPE']:
            if idx < len(row):
                v = row[idx].value
                s = str(v).strip() if v is not None else ''
                if s:
                    if s.isdigit():
                        facility_type_numeric = v
                    else:
                        facility_type_text = v
        facility_no_val = cell(header_indices['FACILITYNO'])

        # Normalize
        abbr = str(abbr_val).strip().upper() if abbr_val else ''
        fac_name = str(name_val).strip() if name_val else ''
        lga_num = _normalize_num_str(lga_number_numeric if lga_number_numeric is not None else lga_number_textual, 2)
        fac_type = _normalize_num_str(facility_type_numeric if facility_type_numeric is not None else facility_type_text, None)
        fac_no = _normalize_num_str(facility_no_val, 4)

        # Require numeric codes to exist
        if not (lga_num and fac_type and fac_no):
            continue

        by_abbr = (tgt_abbr and abbr == tgt_abbr and lga_num == tgt_num and fac_type == tgt_type and fac_no == tgt_fac)
        by_name = (tgt_name and lga_num == tgt_num and fac_type == tgt_type and fac_no == tgt_fac)
        if by_abbr or by_name:
            return fac_name or None

        if not fallback_match_name and (lga_num == tgt_num and fac_type == tgt_type and fac_no == tgt_fac):
            fallback_match_name = fac_name or None

    return fallback_match_name