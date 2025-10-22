from django.contrib import admin
# Removed external RichTextAdmin import; using local definition below
from .models import ImmunizationMaster, ImmunizationSchedule, ImmunizationApproval, VaccinationEventLog, AuditLog, ImmunizationRuleConfig, ImmunizationCertificate
from django.contrib import messages
from django.shortcuts import render, redirect

from django.db import models
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from ckeditor.widgets import CKEditorWidget

from .models import ImmunizationSchedule, ImmunizationMaster


class RichTextAdmin(admin.ModelAdmin):
    formfield_overrides = {models.TextField: {'widget': CKEditorWidget}}


@admin.register(ImmunizationMaster)
class ImmunizationMasterAdmin(RichTextAdmin):
    list_display = ('name', 'interval_value', 'interval_unit', 'is_active')
    list_filter = ('interval_unit', 'is_active')
    search_fields = ('name',)
    change_list_template = 'admin/immunizationmaster_change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('import/', self.admin_site.admin_view(self.import_view), name='immunization_master_import'),
        ]
        return custom + urls

    def import_view(self, request):
        if request.method == 'POST':
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please choose an Excel (.xlsx) file to upload.')
                return redirect('admin:immunization_master_import')

            created, updated, skipped = 0, 0, 0
            try:
                # Try to read via openpyxl
                from openpyxl import load_workbook
                wb = load_workbook(filename=file, data_only=True)
                ws = wb.active
                # Expect headers similar to: Minimum Target Age of Child, Type of Vaccine, Dosage, Route of Administration
                # We'll use Age and Vaccine columns
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                # Find column indexes robustly
                def find_idx(names):
                    for i, h in enumerate(headers):
                        if h and str(h).strip().lower() in names:
                            return i
                    return None
                age_idx = find_idx({'minimum target age of child', 'age', 'recommended age'})
                vaccine_idx = find_idx({'type of vaccine', 'vaccine', 'vaccine name'})
                desc_idx = find_idx({'description', 'notes'})
                dosage_idx = find_idx({'dosage', 'dose'})
                route_idx = find_idx({'route of administration', 'route'})

                if age_idx is None or vaccine_idx is None:
                    messages.error(request, 'Could not detect required columns (Age, Vaccine) in the uploaded sheet.')
                    return redirect('admin:immunization_master_import')

                # Simple parser for age like: "At birth", "6 weeks", "9 months"
                def parse_interval(text):
                    t = str(text or '').strip().lower()
                    if not t:
                        return 0, 'days'
                    if t.startswith('at birth') or t == 'birth':
                        return 0, 'days'
                    parts = t.split()
                    try:
                        val = int(parts[0])
                    except Exception:
                        return 0, 'days'
                    unit = 'days'
                    for w in parts[1:]:
                        if w.startswith('week'):
                            unit = 'weeks'
                            break
                        if w.startswith('month'):
                            unit = 'months'
                            break
                    return val, unit

                from .models import ImmunizationMaster
                for row in ws.iter_rows(min_row=2):
                    age_text = row[age_idx].value if age_idx is not None else ''
                    vaccine_name = row[vaccine_idx].value if vaccine_idx is not None else ''
                    description = row[desc_idx].value if desc_idx is not None else ''
                    dosage = row[dosage_idx].value if dosage_idx is not None else ''
                    route = row[route_idx].value if route_idx is not None else ''
                    if not vaccine_name:
                        skipped += 1
                        continue
                    interval_value, interval_unit = parse_interval(age_text)
                    # Merge dosage/route into description if present
                    extra_parts = []
                    if dosage:
                        extra_parts.append(f"Dosage: {dosage}")
                    if route:
                        extra_parts.append(f"Route: {route}")
                    extra_text = '; '.join(extra_parts)
                    if description and extra_text:
                        description_text = f"{description} ({extra_text})"
                    else:
                        description_text = description or extra_text
                    obj, created_flag = ImmunizationMaster.objects.update_or_create(
                        name=str(vaccine_name).strip(),
                        defaults={
                            'description': description_text or '',
                            'interval_value': max(0, int(interval_value)),
                            'interval_unit': interval_unit,
                            'is_active': True,
                        }
                    )
                    if created_flag:
                        created += 1
                    else:
                        updated += 1

                messages.success(request, f'Import completed: {created} created, {updated} updated, {skipped} skipped.')
                return redirect('admin:immunization_immunizationmaster_changelist')
            except ImportError:
                messages.error(request, 'openpyxl is required to import .xlsx files. Please install it on the server.')
                return redirect('admin:immunization_master_import')
            except Exception as e:
                messages.error(request, f'Import failed: {e}')
                return redirect('admin:immunization_master_import')

        # GET: show upload form
        return render(request, 'admin/immunization_import.html', {
            'title': 'Import Immunization Schedule (Excel)',
        })


@admin.register(ImmunizationSchedule)
class ImmunizationScheduleAdmin(RichTextAdmin):
    list_display = ('baby', 'vaccine_name', 'scheduled_date', 'status', 'date_completed', 'administered_by', 'administered_at', 'visible_to_mother')
    list_filter = ('status', 'scheduled_date', 'date_completed', 'administered_by', 'visible_to_mother')
    search_fields = ('baby__name', 'vaccine_name', 'administered_by__username')


@admin.register(ImmunizationApproval)
class ImmunizationApprovalAdmin(admin.ModelAdmin):
    list_display = ('baby', 'approved_by', 'approved_at')
    search_fields = ('baby__name', 'approved_by__username')


@admin.register(VaccinationEventLog)
class VaccinationEventLogAdmin(admin.ModelAdmin):
    list_display = ('schedule', 'event_type', 'performed_by', 'timestamp')
    list_filter = ('event_type', 'timestamp')
    search_fields = ('schedule__vaccine_name', 'performed_by__username')


@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display = ('model_name', 'object_id', 'action', 'changed_by', 'timestamp')
    list_filter = ('model_name', 'action')
    search_fields = ('model_name', 'object_id', 'changed_by__username')


@admin.register(ImmunizationRuleConfig)
class ImmunizationRuleConfigAdmin(admin.ModelAdmin):
    list_display = ('reschedule_window_days', 'pre_due_reminder_days', 'observation_reminder_hours', 'missed_after_days', 'hospital_name')


@admin.register(ImmunizationCertificate)
class ImmunizationCertificateAdmin(admin.ModelAdmin):
    list_display = ('baby', 'generated_at', 'generated_by')
    search_fields = ('baby__name', 'generated_by__username')
